from __future__ import annotations

import csv
import json
import os
from typing import Iterable

from langchain_core.documents import Document

from rag.document_parsers.base import BaseDocumentParser
from utils.file_handler import txt_loader

try:
    from langchain_text_splitters import MarkdownHeaderTextSplitter
except ImportError:
    from langchain.text_splitter import MarkdownHeaderTextSplitter


def _normalize_cell(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _normalize_headers(header_row: list[str], width: int) -> list[str]:
    headers = []
    for index in range(width):
        raw_header = header_row[index] if index < len(header_row) else ""
        header = raw_header.strip() if raw_header else ""
        headers.append(header or f"column_{index + 1}")
    return headers


def _is_blank_row(row: Iterable[str]) -> bool:
    return not any((cell or "").strip() for cell in row)


def _tabular_row_to_text(headers: list[str], row: list[str]) -> str:
    pairs = []
    for index, header in enumerate(headers):
        value = row[index] if index < len(row) else ""
        value = value.strip()
        if value:
            pairs.append(f"{header}: {value}")
    return "\n".join(pairs)


def _json_scalar_to_text(value) -> str:
    if value is None:
        return "null"
    if isinstance(value, bool):
        return "true" if value else "false"
    return str(value)


def _json_value_to_text(value) -> str:
    if isinstance(value, dict):
        lines = []
        for key, item in value.items():
            if isinstance(item, (dict, list)):
                rendered = json.dumps(item, ensure_ascii=False, sort_keys=True)
            else:
                rendered = _json_scalar_to_text(item)
            lines.append(f"{key}: {rendered}")
        return "\n".join(lines)
    if isinstance(value, list):
        return json.dumps(value, ensure_ascii=False, sort_keys=True)
    return _json_scalar_to_text(value)


class TxtParser(BaseDocumentParser):
    supported_extensions = (".txt",)

    def parse(self, file_path: str) -> list[Document]:
        documents = txt_loader(file_path)
        for doc in documents:
            if not doc.metadata:
                doc.metadata = {}
            doc.metadata.setdefault("source", file_path)
            doc.metadata["file_type"] = "txt"
        return documents


class PdfParser(BaseDocumentParser):
    supported_extensions = (".pdf",)
    max_structured_text_chunk_chars = 2000
    ocr_bitmap_area_threshold = 0.05
    ocr_mode = "auto_fallback"

    def parse(self, file_path: str) -> list[Document]:
        converter = self._build_converter()
        result = converter.convert(file_path)
        return self._docling_result_to_documents(file_path, result)

    def _build_converter(self):
        try:
            from docling.datamodel.base_models import InputFormat
            from docling.datamodel.pipeline_options import OcrAutoOptions, PdfPipelineOptions
            from docling.document_converter import DocumentConverter, PdfFormatOption
        except ImportError as exc:
            raise ImportError("docling is required to parse .pdf files") from exc

        pipeline_options = PdfPipelineOptions(
            do_ocr=True,
            do_table_structure=True,
            ocr_options=OcrAutoOptions(
                force_full_page_ocr=False,
                bitmap_area_threshold=self.ocr_bitmap_area_threshold,
            ),
        )
        return DocumentConverter(
            format_options={
                InputFormat.PDF: PdfFormatOption(pipeline_options=pipeline_options),
            }
        )

    def _docling_result_to_documents(self, file_path: str, conversion_result) -> list[Document]:
        document = conversion_result.document
        parsed_documents = []
        current_section_path: list[str] = []
        current_text_key = None
        current_text_lines: list[str] = []
        table_index = 1

        def flush_text_block():
            nonlocal current_text_key, current_text_lines
            if current_text_key is None:
                current_text_lines = []
                return

            page_no, section_path = current_text_key
            content_lines = [line for line in current_text_lines if line.strip()]
            if content_lines:
                heading_lines = self._markdown_heading_lines(section_path)
                page_content = "\n\n".join(heading_lines + content_lines).strip()
                if page_content:
                    parsed_documents.append(
                        self.create_document(
                            file_path,
                            page_content,
                            **self._pdf_document_metadata(
                                page_no=page_no,
                                section_path=section_path,
                                content_type="text",
                            ),
                        )
                    )
            current_text_key = None
            current_text_lines = []

        for item, level in document.iterate_items():
            label_name = self._item_label_name(item)
            page_no = self._item_page_no(item)

            if label_name == "SECTION_HEADER":
                flush_text_block()
                header_text = self._item_text(item)
                if header_text:
                    current_section_path = self._next_section_path(current_section_path, level, header_text)
                continue

            if hasattr(item, "export_to_markdown"):
                flush_text_block()
                table_markdown = item.export_to_markdown()
                if not table_markdown or not table_markdown.strip():
                    continue
                heading_lines = self._markdown_heading_lines(current_section_path)
                page_content = "\n\n".join(heading_lines + [table_markdown.strip()]).strip()
                parsed_documents.append(
                    self.create_document(
                        file_path,
                        page_content,
                        **self._pdf_document_metadata(
                            page_no=page_no,
                            section_path=current_section_path,
                            content_type="table",
                            table_index=table_index,
                        ),
                    )
                )
                table_index += 1
                continue

            text = self._item_text(item)
            if not text:
                continue

            text_key = (page_no, tuple(current_section_path))
            if current_text_key != text_key:
                flush_text_block()
                current_text_key = text_key
            current_text_lines.append(text)

        flush_text_block()
        return parsed_documents

    @staticmethod
    def _item_label_name(item) -> str:
        label = getattr(item, "label", None)
        return getattr(label, "name", "") or ""

    @staticmethod
    def _item_text(item) -> str:
        return (getattr(item, "text", "") or "").strip()

    @staticmethod
    def _item_page_no(item):
        prov = getattr(item, "prov", None) or []
        if prov:
            return getattr(prov[0], "page_no", None)
        return None

    @staticmethod
    def _next_section_path(current_path: list[str], level: int, header_text: str) -> list[str]:
        normalized_level = max(1, min(level or 1, 6))
        return [*current_path[: normalized_level - 1], header_text]

    @staticmethod
    def _markdown_heading_lines(section_path: tuple[str, ...] | list[str]) -> list[str]:
        lines = []
        for index, title in enumerate(section_path, start=1):
            if title:
                lines.append(f"{'#' * min(index, 6)} {title}")
        return lines

    def _pdf_document_metadata(
        self,
        page_no,
        section_path: tuple[str, ...] | list[str],
        content_type: str,
        table_index: int | None = None,
    ) -> dict:
        metadata = {
            "file_type": "pdf",
            "page": page_no,
            "section_path": " > ".join(section_path),
            "section_title": section_path[-1] if section_path else "",
            "content_type": content_type,
            "ocr_enabled": True,
            "ocr_mode": self.ocr_mode,
            "ocr_bitmap_area_threshold": self.ocr_bitmap_area_threshold,
        }
        if table_index is not None:
            metadata["table_index"] = table_index
        return metadata


class CsvParser(BaseDocumentParser):
    supported_extensions = (".csv",)

    def parse(self, file_path: str) -> list[Document]:
        with open(file_path, "r", encoding="utf-8-sig", newline="") as f:
            rows = [[_normalize_cell(cell) for cell in row] for row in csv.reader(f)]

        if not rows:
            return []

        width = max((len(row) for row in rows), default=0)
        if width == 0:
            return []

        headers = _normalize_headers(rows[0], width)
        documents = []
        for row_number, row in enumerate(rows[1:], start=2):
            normalized_row = row + [""] * (width - len(row))
            if _is_blank_row(normalized_row):
                continue
            page_content = _tabular_row_to_text(headers, normalized_row)
            if not page_content:
                continue
            documents.append(
                self.create_document(
                    file_path,
                    page_content,
                    file_type="csv",
                    row_index=row_number,
                )
            )
        return documents


class XlsxParser(BaseDocumentParser):
    supported_extensions = (".xlsx",)

    def parse(self, file_path: str) -> list[Document]:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise ImportError("openpyxl is required to parse .xlsx files") from exc

        workbook = load_workbook(filename=file_path, data_only=True, read_only=True)
        documents = []
        try:
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                rows = [
                    [_normalize_cell(cell) for cell in row]
                    for row in sheet.iter_rows(values_only=True)
                ]
                if not rows:
                    continue

                width = max((len(row) for row in rows), default=0)
                if width == 0:
                    continue

                headers = _normalize_headers(rows[0], width)
                for row_number, row in enumerate(rows[1:], start=2):
                    normalized_row = row + [""] * (width - len(row))
                    if _is_blank_row(normalized_row):
                        continue
                    page_content = _tabular_row_to_text(headers, normalized_row)
                    if not page_content:
                        continue
                    documents.append(
                        self.create_document(
                            file_path,
                            page_content,
                            file_type="xlsx",
                            sheet_name=sheet_name,
                            row_index=row_number,
                        )
                    )
        finally:
            workbook.close()
        return documents


class JsonParser(BaseDocumentParser):
    supported_extensions = (".json",)

    def parse(self, file_path: str) -> list[Document]:
        with open(file_path, "r", encoding="utf-8") as f:
            payload = json.load(f)

        if isinstance(payload, list):
            documents = []
            for index, item in enumerate(payload):
                page_content = _json_value_to_text(item)
                if not page_content.strip():
                    continue
                documents.append(
                    self.create_document(
                        file_path,
                        page_content,
                        file_type="json",
                        record_index=index,
                        json_path=f"$[{index}]",
                    )
                )
            return documents

        if isinstance(payload, dict):
            documents = []
            for key, value in payload.items():
                if isinstance(value, (dict, list)):
                    rendered_value = json.dumps(value, ensure_ascii=False, sort_keys=True, indent=2)
                    page_content = f"{key}:\n{rendered_value}"
                else:
                    page_content = f"{key}: {_json_scalar_to_text(value)}"
                if not page_content.strip():
                    continue
                documents.append(
                    self.create_document(
                        file_path,
                        page_content,
                        file_type="json",
                        json_path=f"$.{key}",
                    )
                )
            return documents

        return [
            self.create_document(
                file_path,
                _json_scalar_to_text(payload),
                file_type="json",
                json_path="$",
            )
        ]


class MarkdownParser(BaseDocumentParser):
    supported_extensions = (".md",)

    headers_to_split_on = [
        ("#", "h1"),
        ("##", "h2"),
        ("###", "h3"),
        ("####", "h4"),
        ("#####", "h5"),
        ("######", "h6"),
    ]

    def parse(self, file_path: str) -> list[Document]:
        with open(file_path, "r", encoding="utf-8") as f:
            text = f.read()

        if not text.strip():
            return []

        splitter = MarkdownHeaderTextSplitter(
            headers_to_split_on=self.headers_to_split_on,
            strip_headers=False,
        )
        documents = splitter.split_text(text)
        if not documents:
            return [self.create_document(file_path, text, file_type="md")]

        parsed_documents = []
        for doc in documents:
            metadata = dict(doc.metadata or {})
            header_path_parts = [metadata.get(level) for _, level in self.headers_to_split_on if metadata.get(level)]
            header_path = " > ".join(header_path_parts)
            page_content = doc.page_content.strip()
            if header_path and header_path not in page_content:
                page_content = f"{header_path}\n\n{page_content}"
            if not page_content:
                continue
            parsed_documents.append(
                self.create_document(
                    file_path,
                    page_content,
                    file_type="md",
                    markdown_path=header_path,
                    **metadata,
                )
            )
        return parsed_documents


class DocxParser(BaseDocumentParser):
    supported_extensions = (".docx",)

    def parse(self, file_path: str) -> list[Document]:
        try:
            from docx import Document as DocxDocument
            from docx.document import Document as DocxDocumentType
            from docx.oxml.table import CT_Tbl
            from docx.oxml.text.paragraph import CT_P
            from docx.table import Table
            from docx.text.paragraph import Paragraph
        except ImportError as exc:
            raise ImportError("python-docx is required to parse .docx files") from exc

        def iter_block_items(parent: DocxDocumentType):
            for child in parent.element.body.iterchildren():
                if isinstance(child, CT_P):
                    yield Paragraph(child, parent)
                elif isinstance(child, CT_Tbl):
                    yield Table(child, parent)

        document = DocxDocument(file_path)
        default_title = os.path.splitext(os.path.basename(file_path))[0]
        current_title = default_title
        current_lines: list[str] = []
        parsed_documents = []

        def flush_current_section():
            nonlocal current_lines
            body = "\n\n".join(line for line in current_lines if line.strip()).strip()
            if not body:
                current_lines = []
                return
            page_content = body if current_title == default_title else f"{current_title}\n\n{body}"
            parsed_documents.append(
                self.create_document(
                    file_path,
                    page_content,
                    file_type="docx",
                    section_title=current_title,
                )
            )
            current_lines = []

        for block in iter_block_items(document):
            if isinstance(block, Paragraph):
                text = block.text.strip()
                if not text:
                    continue
                style_name = block.style.name if block.style else ""
                if style_name.startswith("Heading"):
                    flush_current_section()
                    current_title = text
                    continue
                current_lines.append(text)
                continue

            table_text = self._table_to_text(block)
            if table_text:
                current_lines.append(table_text)

        flush_current_section()
        return parsed_documents

    @staticmethod
    def _table_to_text(table) -> str:
        rows = []
        for row in table.rows:
            values = [_normalize_cell(cell.text) for cell in row.cells]
            if _is_blank_row(values):
                continue
            rows.append(values)

        if not rows:
            return ""

        width = max((len(row) for row in rows), default=0)
        headers = _normalize_headers(rows[0], width)
        rendered_rows = []
        for row in rows[1:]:
            normalized_row = row + [""] * (width - len(row))
            row_text = _tabular_row_to_text(headers, normalized_row)
            if row_text:
                rendered_rows.append(row_text)

        if rendered_rows:
            return "\n\n".join(rendered_rows)

        fallback = rows[0] + [""] * (width - len(rows[0]))
        return _tabular_row_to_text(headers, fallback)
